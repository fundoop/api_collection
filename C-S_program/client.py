from tkinter import *
from tkinter.filedialog import *
import socket
import json
import socket, threading
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import time
def get_sql(orgname):
    sql = """sql
    WHERE
    a.orgcode in """+orgname+"""
    GROUP BY
    a.orgcode,a.orgname"""
    sql1="""
    sql
    a.orgcode in """+orgname+"""
    GROUP BY
    a.orgcode,a.orgname,
    DATE_FORMAT(a.date,'%Y-%m')
    """
    sql2="""
    sql
    WHERE orgcode in """+orgname
    sql3="""sql
    where orgcode in """+orgname+""")
    and   a.ordersn=b.ordersn
    and   a.orgcode=c.orgcode"""
    sql4="""sql
    where a.orgcode in """+orgname+""") p
    group by left(data_date,6)"""
    sql5="""sql
    where a.orgcode in """+orgname
    return [sql,sql1,sql2,sql3,sql4,sql5]
class WidgetsDemo:
    def __init__(self):
        self.opt=[]
        self.org=''
        self.orgcode=''
        window = Tk()
        #window.geometry('80x80+10+10')
        window.title("风控数据")
        #添加一个多选按钮和单选按钮到frame1
        #frame1 = Frame(window)
        #frame1.pack()  #看下面的解释（包管理器）
        frame2 = Frame(window)
        #frame2.pack()
        frame2.grid(row = 0, column = 0,sticky = W)
        label1 = Label(frame2, text = "请输入机构名称：",width=14)
        label1.grid(row = 0, column = 0,sticky = W)
        label2 = Label(frame2, text = "选项：",width=14)
        label2.grid(row = 2, column = 0,sticky = W)
        self.name = StringVar()
        self.options=StringVar()
        entryName = Entry(frame2, textvariable = self.name)
        entryName.grid(row = 0, column = 1,sticky = W)
        entryName.bind('<Return>',self.processButton)
        label1 = Label(frame2, text = "-----------------------------------------------------------")
        label1.grid(row = 1, column = 0,columnspan=3)
        self.chose = Entry(frame2, textvariable = self.options,width=7)
        self.chose.grid(row = 2, column = 1,sticky = W)        
        self.chose.bind('<Return>',self.choose)
        btGetName = Button(frame2, text = "提交", command = self.processButton,width=7)
        btGetName.grid(row = 0, column = 2,sticky = W)
        btGetName = Button(frame2, text = "清除", command = self.clearButton,width=7)
        btGetName.grid(row = 2, column = 2,sticky = W)
        self.text = Text(window,width=45)
        self.text.grid(row = 1, column = 0)
        self.text.insert(1.0, "开始----------------------------------------\n")
        window.mainloop()
    def clearButton(self):
        self.text.delete(1.0,END)
        self.options.set('')
        self.name.set('')
    def get_xlsx(self,data,book,sheet):
        try:
            wb = load_workbook(r'./'+book+'.xlsx')
        except:wb=Workbook()
        try:
            wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
        except:pass
        try:
            ws = wb.create_sheet(sheet)
        except:self.text.insert(1.0,'工作表名不能重复')
        try:
            rows = len(data)
            cols = len(data[0])
            i=1
            for rx in range(rows):
                  for cx in range(cols):
                        ws.cell(row=rx+1, column=cx+1).value = data[rx][cx]
            wb.save(filename=r'./'+book+'.xlsx')
            self.text.insert(1.0,'\n'+book+'.'+sheet+'(收到'+str(rows-1)+'条数据)')
        except:
            wb.save(filename=r'./'+book+'.xlsx')
            self.text.insert(1.0,'\n'+book+'.'+sheet+'(收到0条数据)')
            return 0
    def get_data(self,request):
        s=socket.socket(socket.AF_INET,socket.SOCK_STREAM)
        s.setsockopt(socket.SOL_SOCKET,socket.SO_REUSEADDR,1)
        addr=('172.23.129.244',140)
        s.connect(addr)
        s.send(json.dumps(request).encode())
        data=[]
        data_count=-1
        while True:
            response=json.loads(s.recv(4096).decode())
            if response[0]=='提示':
                self.text.insert(1.0,'\n'+response[1])
            elif response[0]=='数据条数':
                data_count=int(response[1])
                s.send(json.dumps(['条数确认','']).encode())
            elif response[0]=='数据':
                if len(data)==0:data=response[1]
                else:data=data+response[1]
                s.send(json.dumps(['确认','']).encode())
            if len(data)==data_count:
                break
        s.close()
        return data
    def choose(self,lock=1):
        try:
            if self.opt==[] or self.org=='':return
            choice=self.options.get()
            print(choice)
            cnt=len(self.opt)
            if choice==str(len(self.opt))+'、以上所有':
                time.sleep(3)
                self.choose()
            choice=int(choice)
            t=()
            if choice==cnt:
                for i in range(1,cnt):
                    t=t+(self.opt[i][0],)
                self.orgcode=t
                self.text.insert(1.0,'\n已选择：所有机构')
            elif choice in range(1,cnt):
                t="('"+self.opt[choice][0]+"')"
                choice=self.opt[choice][1]
                self.orgcode=t
                self.org=choice
                self.text.insert(1.0,'\n已选择：'+choice)
            else:
                self.text.insert(1.0,'\n请按照数字重新选择')
        except:self.text.insert(1.0,'\n请按照数字重新选择')

    def main(self):
        #try:
        sql="""SELECT orgcode,orgname
        FROM link_etcaccounts
        WHERE orgname LIKE '%"""+self.org+"""%'"""
        self.opt=self.get_data(['数据请求',self.org,'机构选项',sql,'cat_link'])
        if self.opt==[] or len(self.opt)==0:
            self.text.insert(1.0, '\n无类似机构')
            return 0
        elif len(self.opt)==2:
            self.text.insert(1.0,'\n已开始取数，请确认：'+str(self.opt[1][1]))
            self.orgcode="('"+self.opt[1][0]+"')"
        elif len(self.opt)>2:
            bb='以下类似机构，请选择：'
            for i in range(1,len(self.opt)):bb=bb+'\n'+str(i)+'、'+str(self.opt[i][1])
            bb=bb+'\n'+str(len(self.opt))+'、以上所有\n'
            self.text.insert(1.0,'\n'+bb)
            self.text.insert(1.0,'\n10秒后默认取所有机构\n')
            for i in range(1,10):
                if self.orgcode=='':
                    time.sleep(1)
                    self.text.insert(1.0,'\n'+str(10-i)+'秒后默认取所有机构\n')
            
            if self.orgcode=='':
                t=()
                for i in range(1,len(self.opt)):
                    t=t+(self.opt[i][0],)
                self.orgcode=t
            print(self.orgcode)
            books=['机构概况','机构月消费','机构日消费','消费明细','设备汇总','设备明细']
            dbs=['cat_link','cat_link','cat_link','cat_link','oms','oms']
            sqls=get_sql(str(self.orgcode))
            self.orgcode=''
            for i in range(0,6):
                request=['数据请求',self.org,books[i],sqls[i],dbs[i]]
                self.get_xlsx(self.get_data(request),request[1],request[2])
                self.text.insert(1.0,'\n----------------------------------------')
            self.text.insert(1.0,'\n已结束，请查看同文件夹下文件')
            return 0
        else:
            self.text.insert(1.0,'\n未知错误')
            return 0

        books=['机构概况','机构月消费','机构日消费','消费明细','设备汇总','设备明细']
        dbs=['cat_link','cat_link','cat_link','cat_link','oms','oms']
        sqls=get_sql(str(self.orgcode))
        for i in range(0,6):
            request=['数据请求',self.org,books[i],sqls[i],dbs[i]]
            self.get_xlsx(self.get_data(request),request[1],request[2])
            self.text.insert(1.0,'\n----------------------------------------')
        self.text.insert(1.0,'\n已结束，请查看同文件夹下文件')
        #except Exception as e:self.text.insert(END,'\n'+str(e)+'\n')
    def processButton(self,t=1):
        self.org=self.name.get()
        if len(self.org)==0:
            self.text.insert(1.0,'\n请先输入机构名')
            return 0
        t = threading.Thread(target = self.main)
        # 线程运行
        t.start()
WidgetsDemo()
